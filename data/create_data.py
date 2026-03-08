"""Script to generate the Customer Service Tables.xlsx file with fake data."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta
import random

random.seed(42)

wb = openpyxl.Workbook()


# ── Helpers ──────────────────────────────────────────────────────────────────

def style_header(ws, row: int = 1) -> None:
    """
    Apply bold + blue header style to the first row of a worksheet.

    param ws: The worksheet to style.
    param row: The row number to apply header style to.
    """
    fill = PatternFill("solid", fgColor="4472C4")
    for cell in ws[row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")


def rand_date(start: datetime, end: datetime) -> datetime:
    """
    Generate a random datetime between start and end.

    param start: Start of the date range.
    param end: End of the date range.
    """
    delta = end - start
    return start + timedelta(seconds=random.randint(0, int(delta.total_seconds())))


# ── Sheet 1: Table Metadata ───────────────────────────────────────────────────

ws_meta = wb.active
ws_meta.title = "Table Metadata"

meta_headers = ["Table Name", "Column Name", "Data Type", "Nullable", "Description", "Relationships"]
ws_meta.append(meta_headers)

metadata = [
    ("customers_tbl", "customer_id",   "INTEGER",  "NO",  "Unique customer identifier",                   "PK"),
    ("customers_tbl", "name",           "VARCHAR",  "NO",  "Full name of the customer",                    ""),
    ("customers_tbl", "email",          "VARCHAR",  "NO",  "Email address of the customer",                ""),
    ("customers_tbl", "tier",           "VARCHAR",  "NO",  "Customer tier: standard, premium, enterprise", ""),
    ("customers_tbl", "region",         "VARCHAR",  "YES", "Geographic region of the customer",            ""),
    ("customers_tbl", "created_at",     "DATETIME", "NO",  "Date the customer account was created",        ""),

    ("agents_tbl",    "agent_id",       "INTEGER",  "NO",  "Unique agent identifier",                      "PK"),
    ("agents_tbl",    "name",           "VARCHAR",  "NO",  "Full name of the agent",                       ""),
    ("agents_tbl",    "email",          "VARCHAR",  "NO",  "Work email of the agent",                      ""),
    ("agents_tbl",    "team",           "VARCHAR",  "NO",  "Team the agent belongs to: billing, tech, general", ""),
    ("agents_tbl",    "hire_date",      "DATE",     "NO",  "Date the agent was hired",                     ""),
    ("agents_tbl",    "is_active",      "BOOLEAN",  "NO",  "Whether the agent is currently active",        ""),

    ("categories_tbl","category_id",    "INTEGER",  "NO",  "Unique category identifier",                   "PK"),
    ("categories_tbl","name",           "VARCHAR",  "NO",  "Category name (e.g. Billing, Technical)",      ""),
    ("categories_tbl","description",    "VARCHAR",  "YES", "Description of the category",                  ""),
    ("categories_tbl","sla_hours",      "INTEGER",  "NO",  "SLA resolution target in hours",               ""),

    ("tickets_tbl",   "ticket_id",      "INTEGER",  "NO",  "Unique ticket identifier",                     "PK"),
    ("tickets_tbl",   "customer_id",    "INTEGER",  "NO",  "Customer who raised the ticket",               "FK -> customers_tbl.customer_id"),
    ("tickets_tbl",   "agent_id",       "INTEGER",  "YES", "Agent assigned to the ticket (NULL if unassigned)", "FK -> agents_tbl.agent_id"),
    ("tickets_tbl",   "category_id",    "INTEGER",  "NO",  "Category of the ticket",                       "FK -> categories_tbl.category_id"),
    ("tickets_tbl",   "status",         "VARCHAR",  "NO",  "Ticket status: open, pending, resolved",       ""),
    ("tickets_tbl",   "priority",       "VARCHAR",  "NO",  "Priority: low, medium, high, critical",        ""),
    ("tickets_tbl",   "subject",        "VARCHAR",  "NO",  "Short description of the issue",               ""),
    ("tickets_tbl",   "created_at",     "DATETIME", "NO",  "When the ticket was created",                  ""),
    ("tickets_tbl",   "resolved_at",    "DATETIME", "YES", "When the ticket was resolved (NULL if open)",  ""),

    ("interactions_tbl","interaction_id","INTEGER", "NO",  "Unique interaction identifier",                "PK"),
    ("interactions_tbl","ticket_id",    "INTEGER",  "NO",  "Ticket this interaction belongs to",           "FK -> tickets_tbl.ticket_id"),
    ("interactions_tbl","agent_id",     "INTEGER",  "NO",  "Agent who performed the interaction",          "FK -> agents_tbl.agent_id"),
    ("interactions_tbl","type",         "VARCHAR",  "NO",  "Interaction type: call, email, chat",          ""),
    ("interactions_tbl","notes",        "TEXT",     "YES", "Notes from the interaction",                   ""),
    ("interactions_tbl","created_at",   "DATETIME", "NO",  "When the interaction occurred",                ""),

    ("products_tbl",  "product_id",     "INTEGER",  "NO",  "Unique product identifier",                    "PK"),
    ("products_tbl",  "name",           "VARCHAR",  "NO",  "Product name",                                 ""),
    ("products_tbl",  "version",        "VARCHAR",  "YES", "Product version string",                       ""),
    ("products_tbl",  "category",       "VARCHAR",  "NO",  "Product category: SaaS, Hardware, API",        ""),

    ("ticket_products_tbl","ticket_id", "INTEGER",  "NO",  "Ticket linked to the product",                 "FK -> tickets_tbl.ticket_id"),
    ("ticket_products_tbl","product_id","INTEGER",  "NO",  "Product linked to the ticket",                 "FK -> products_tbl.product_id"),

    ("feedback_tbl",  "feedback_id",    "INTEGER",  "NO",  "Unique feedback identifier",                   "PK"),
    ("feedback_tbl",  "ticket_id",      "INTEGER",  "NO",  "Ticket the feedback is about",                 "FK -> tickets_tbl.ticket_id"),
    ("feedback_tbl",  "customer_id",    "INTEGER",  "NO",  "Customer who submitted the feedback",          "FK -> customers_tbl.customer_id"),
    ("feedback_tbl",  "rating",         "INTEGER",  "NO",  "CSAT score: 1 (worst) to 5 (best)",            ""),
    ("feedback_tbl",  "comment",        "TEXT",     "YES", "Optional comment from the customer",           ""),
    ("feedback_tbl",  "submitted_at",   "DATETIME", "NO",  "When the feedback was submitted",              ""),

    ("escalations_tbl","escalation_id", "INTEGER",  "NO",  "Unique escalation identifier",                 "PK"),
    ("escalations_tbl","ticket_id",     "INTEGER",  "NO",  "Ticket that was escalated",                    "FK -> tickets_tbl.ticket_id"),
    ("escalations_tbl","from_agent_id", "INTEGER",  "NO",  "Agent who escalated the ticket",               "FK -> agents_tbl.agent_id"),
    ("escalations_tbl","to_agent_id",   "INTEGER",  "NO",  "Agent who received the escalation",            "FK -> agents_tbl.agent_id"),
    ("escalations_tbl","reason",        "TEXT",     "YES", "Reason for escalation",                        ""),
    ("escalations_tbl","created_at",    "DATETIME", "NO",  "When the escalation occurred",                 ""),
]

for row in metadata:
    ws_meta.append(row)

style_header(ws_meta)
for col in ws_meta.columns:
    ws_meta.column_dimensions[col[0].column_letter].width = 28


# ── Sheet 2: Scenarios ────────────────────────────────────────────────────────

ws_scenarios = wb.create_sheet("Scenarios")
scenario_headers = ["#", "Scenario", "Tables Involved", "Complexity", "Expected Output"]
ws_scenarios.append(scenario_headers)

scenarios = [
    (1,  "List all open tickets with high or critical priority for premium or enterprise customers, including the customer name and agent name.",
         "tickets_tbl, customers_tbl, agents_tbl", "Medium",
         "ticket_id, customer_name, tier, agent_name, priority, created_at"),
    (2,  "Find all agents who currently have more than 3 open tickets assigned to them.",
         "tickets_tbl, agents_tbl", "Medium",
         "agent_id, agent_name, open_ticket_count"),
    (3,  "List customers who have raised tickets in more than 2 different categories.",
         "tickets_tbl, customers_tbl, categories_tbl", "Medium",
         "customer_id, customer_name, category_count"),
    (4,  "Calculate the average resolution time (in hours) per agent for resolved tickets, ordered by fastest average.",
         "tickets_tbl, agents_tbl", "Medium",
         "agent_id, agent_name, avg_resolution_hours"),
    (5,  "Find all tickets that have breached their SLA (open tickets where created_at is older than the category sla_hours threshold).",
         "tickets_tbl, categories_tbl", "Hard",
         "ticket_id, subject, category_name, sla_hours, hours_open"),
    (6,  "Show the top 3 categories by number of unresolved tickets this month.",
         "tickets_tbl, categories_tbl", "Medium",
         "category_name, open_ticket_count"),
    (7,  "List all premium or enterprise customers who have an open ticket with no agent assigned.",
         "tickets_tbl, customers_tbl", "Easy",
         "customer_id, customer_name, tier, ticket_id, created_at"),
    (8,  "For each agent, show the number of interactions they have handled broken down by interaction type (call, email, chat).",
         "interactions_tbl, agents_tbl", "Hard",
         "agent_name, call_count, email_count, chat_count"),
    (9,  "Find customers who have had more than one ticket in the same category.",
         "tickets_tbl, customers_tbl, categories_tbl", "Hard",
         "customer_name, category_name, ticket_count"),
    (10, "Show the monthly ticket volume trend for the last 6 months grouped by status.",
         "tickets_tbl", "Medium",
         "month, status, ticket_count"),
    (11, "Find all products that have more than 2 open or pending tickets linked to them.",
         "tickets_tbl, ticket_products_tbl, products_tbl", "Medium",
         "product_id, product_name, open_ticket_count"),
    (12, "Show the average CSAT rating per agent for resolved tickets, only for agents with at least 3 feedback entries.",
         "feedback_tbl, tickets_tbl, agents_tbl", "Hard",
         "agent_name, avg_rating, feedback_count"),
    (13, "List all tickets that were escalated more than once.",
         "escalations_tbl, tickets_tbl", "Medium",
         "ticket_id, subject, escalation_count"),
    (14, "Find agents who received escalations but have an average CSAT rating below 3.",
         "escalations_tbl, feedback_tbl, tickets_tbl, agents_tbl", "Hard",
         "agent_name, escalations_received, avg_csat"),
    (15, "Show which product category has the highest average resolution time.",
         "tickets_tbl, ticket_products_tbl, products_tbl", "Hard",
         "product_category, avg_resolution_hours"),
]

for row in scenarios:
    ws_scenarios.append(row)

style_header(ws_scenarios)
ws_scenarios.column_dimensions["A"].width = 5
ws_scenarios.column_dimensions["B"].width = 70
ws_scenarios.column_dimensions["C"].width = 40
ws_scenarios.column_dimensions["D"].width = 12
ws_scenarios.column_dimensions["E"].width = 45


# ── Sheet 3: customers_tbl ────────────────────────────────────────────────────

ws_customers = wb.create_sheet("customers_tbl")
ws_customers.append(["customer_id", "name", "email", "tier", "region", "created_at"])

tiers = ["standard", "standard", "standard", "premium", "premium", "enterprise"]
regions = ["North", "South", "East", "West", "Central"]
first_names = ["Alice", "Bob", "Carol", "David", "Eve", "Frank", "Grace", "Hank", "Iris", "Jack",
               "Karen", "Leo", "Mia", "Noah", "Olivia", "Pete", "Quinn", "Rose", "Sam", "Tina"]
last_names = ["Smith", "Johnson", "Williams", "Brown", "Jones", "Garcia", "Miller", "Davis",
              "Wilson", "Moore", "Taylor", "Anderson", "Thomas", "Jackson", "White"]

start = datetime(2022, 1, 1)
end = datetime(2024, 6, 1)

for i in range(1, 31):
    name = f"{random.choice(first_names)} {random.choice(last_names)}"
    email = name.lower().replace(" ", ".") + f"@example.com"
    ws_customers.append([
        i, name, email,
        random.choice(tiers),
        random.choice(regions),
        rand_date(start, end).strftime("%Y-%m-%d %H:%M:%S"),
    ])

style_header(ws_customers)


# ── Sheet 4: agents_tbl ───────────────────────────────────────────────────────

ws_agents = wb.create_sheet("agents_tbl")
ws_agents.append(["agent_id", "name", "email", "team", "hire_date", "is_active"])

teams = ["billing", "tech", "general"]
agent_names = ["Amy Chen", "Brian Roy", "Carlos Diaz", "Diana Park", "Ethan Brooks",
               "Fiona Hall", "George Kim", "Hannah Wu", "Ian Foster", "Julia Ng"]

for i, name in enumerate(agent_names, 1):
    ws_agents.append([
        i, name, name.lower().replace(" ", ".") + "@company.com",
        random.choice(teams),
        rand_date(datetime(2020, 1, 1), datetime(2023, 12, 31)).strftime("%Y-%m-%d"),
        True if i <= 8 else False,
    ])

style_header(ws_agents)


# ── Sheet 5: categories_tbl ───────────────────────────────────────────────────

ws_categories = wb.create_sheet("categories_tbl")
ws_categories.append(["category_id", "name", "description", "sla_hours"])

categories = [
    (1, "Billing",       "Payment, invoice, and subscription issues", 24),
    (2, "Technical",     "Product bugs, outages, and technical errors", 8),
    (3, "Account",       "Login, password reset, and account access issues", 12),
    (4, "General",       "General inquiries and feedback", 48),
    (5, "Compliance",    "Legal, GDPR, and compliance-related requests", 72),
]

for row in categories:
    ws_categories.append(row)

style_header(ws_categories)
cat_ids = [c[0] for c in categories]


# ── Sheet 6: tickets_tbl ──────────────────────────────────────────────────────

ws_tickets = wb.create_sheet("tickets_tbl")
ws_tickets.append(["ticket_id", "customer_id", "agent_id", "category_id", "status",
                   "priority", "subject", "created_at", "resolved_at"])

statuses = ["open", "open", "pending", "resolved", "resolved"]
priorities = ["low", "medium", "medium", "high", "critical"]
subjects = [
    "Cannot access my account", "Invoice amount is incorrect", "Product keeps crashing",
    "Request for data export", "Password reset not working", "Overcharged on last bill",
    "Feature not working as expected", "Need to update billing info", "Compliance documentation request",
    "Performance degradation noticed", "Error 500 on dashboard", "Refund request",
    "Account locked out", "Missing emails from system", "API rate limit issues",
]

ticket_start = datetime(2024, 1, 1)
ticket_end = datetime(2024, 12, 31)

for i in range(1, 61):
    created = rand_date(ticket_start, ticket_end)
    status = random.choice(statuses)
    resolved = rand_date(created, created + timedelta(hours=random.randint(1, 120))) if status == "resolved" else None
    agent = random.choice(range(1, 11)) if random.random() > 0.15 else None  # 15% unassigned.
    ws_tickets.append([
        i,
        random.randint(1, 30),
        agent,
        random.choice(cat_ids),
        status,
        random.choice(priorities),
        random.choice(subjects),
        created.strftime("%Y-%m-%d %H:%M:%S"),
        resolved.strftime("%Y-%m-%d %H:%M:%S") if resolved else None,
    ])

style_header(ws_tickets)


# ── Sheet 7: interactions_tbl ─────────────────────────────────────────────────

ws_interactions = wb.create_sheet("interactions_tbl")
ws_interactions.append(["interaction_id", "ticket_id", "agent_id", "type", "notes", "created_at"])

interaction_types = ["call", "email", "chat"]
notes_templates = [
    "Reviewed the issue with the customer and provided next steps.",
    "Sent follow-up email with resolution details.",
    "Customer confirmed the issue is resolved.",
    "Escalated to tier 2 support for further investigation.",
    "Waiting on customer to provide additional information.",
]

iid = 1
for ticket_id in range(1, 61):
    num_interactions = random.randint(1, 4)
    t_created = datetime(2024, random.randint(1, 12), random.randint(1, 28))
    for _ in range(num_interactions):
        ws_interactions.append([
            iid,
            ticket_id,
            random.randint(1, 10),
            random.choice(interaction_types),
            random.choice(notes_templates),
            rand_date(t_created, t_created + timedelta(days=3)).strftime("%Y-%m-%d %H:%M:%S"),
        ])
        iid += 1

style_header(ws_interactions)


# ── Sheet 8: products_tbl ────────────────────────────────────────────────────

ws_products = wb.create_sheet("products_tbl")
ws_products.append(["product_id", "name", "version", "category"])

products = [
    (1, "BillingPro",     "3.2.1", "SaaS"),
    (2, "SupportDesk",    "2.0.0", "SaaS"),
    (3, "DataSync API",   "1.4.0", "API"),
    (4, "CloudRouter",    "5.1.2", "API"),
    (5, "OfficeTerminal", "8.0",   "Hardware"),
    (6, "SecureVault",    "2.3.0", "SaaS"),
    (7, "ReportEngine",   "1.1.0", "API"),
]

for row in products:
    ws_products.append(row)

style_header(ws_products)
product_ids = [p[0] for p in products]


# ── Sheet 9: ticket_products_tbl ──────────────────────────────────────────────

ws_tp = wb.create_sheet("ticket_products_tbl")
ws_tp.append(["ticket_id", "product_id"])

linked = set()
for ticket_id in range(1, 61):
    num_products = random.randint(1, 2)
    for pid in random.sample(product_ids, num_products):
        if (ticket_id, pid) not in linked:
            ws_tp.append([ticket_id, pid])
            linked.add((ticket_id, pid))

style_header(ws_tp)


# ── Sheet 10: feedback_tbl ────────────────────────────────────────────────────

ws_feedback = wb.create_sheet("feedback_tbl")
ws_feedback.append(["feedback_id", "ticket_id", "customer_id", "rating", "comment", "submitted_at"])

comments = [
    "Great support, issue resolved quickly.",
    "Took too long but eventually resolved.",
    "Agent was helpful and professional.",
    "Not satisfied, problem still persists.",
    "Average experience, nothing special.",
    None,
]

fid = 1
# Only resolved tickets get feedback (~80% of them).
for ticket_id in range(1, 61):
    if random.random() < 0.6:
        ws_feedback.append([
            fid,
            ticket_id,
            random.randint(1, 30),
            random.randint(1, 5),
            random.choice(comments),
            rand_date(datetime(2024, 6, 1), datetime(2025, 1, 1)).strftime("%Y-%m-%d %H:%M:%S"),
        ])
        fid += 1

style_header(ws_feedback)


# ── Sheet 11: escalations_tbl ────────────────────────────────────────────────

ws_escalations = wb.create_sheet("escalations_tbl")
ws_escalations.append(["escalation_id", "ticket_id", "from_agent_id", "to_agent_id", "reason", "created_at"])

reasons = [
    "Customer requested senior agent.",
    "Issue beyond agent's technical expertise.",
    "Customer dissatisfied with resolution.",
    "Policy exception required.",
    None,
]

eid = 1
# ~30% of tickets get escalated; some get escalated twice.
for ticket_id in range(1, 61):
    if random.random() < 0.30:
        from_agent = random.randint(1, 10)
        to_agent = random.choice([a for a in range(1, 11) if a != from_agent])
        ws_escalations.append([
            eid, ticket_id, from_agent, to_agent,
            random.choice(reasons),
            rand_date(datetime(2024, 1, 1), datetime(2024, 12, 31)).strftime("%Y-%m-%d %H:%M:%S"),
        ])
        eid += 1
        # ~20% chance of a second escalation.
        if random.random() < 0.20:
            from_agent2 = to_agent
            to_agent2 = random.choice([a for a in range(1, 11) if a != from_agent2])
            ws_escalations.append([
                eid, ticket_id, from_agent2, to_agent2,
                random.choice(reasons),
                rand_date(datetime(2024, 1, 1), datetime(2024, 12, 31)).strftime("%Y-%m-%d %H:%M:%S"),
            ])
            eid += 1

style_header(ws_escalations)


# ── Save ──────────────────────────────────────────────────────────────────────

output_path = "data/Customer Service Tables.xlsx"
wb.save(output_path)
print(f"Created: {output_path}")
print(f"Sheets: {wb.sheetnames}")
